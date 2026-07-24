#!/usr/bin/env python3
"""
Export captured data to Excel workbooks that download from the dashboard.

  fares.xlsx     - every captured flight (raw) + the curated slot view
  aviation.xlsx  - DGCA / civil-aviation monthly metrics (if collected)

Both are regenerated on each capture run so the download is always current.
"""

import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import config

FARES_XLSX = "fares.xlsx"
AVIATION_XLSX = "aviation.xlsx"
AVIATION_DB = "aviation.db"


def _autosize(ws, max_width=42):
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 2, max_width)


def _header(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"


def build_fares_excel(path=FARES_XLSX):
    conn = sqlite3.connect(config.DB_PATH)
    conn.row_factory = sqlite3.Row
    wb = Workbook()

    # Sheet 1: every captured flight (the full longitudinal raw dataset)
    ws = wb.active
    ws.title = "All fares"
    cols = ["captured_at", "capture_date", "sector", "origin", "destination",
            "target_date", "horizon", "airline", "flight_number", "dep_time",
            "arr_time", "duration_min", "stops", "price", "currency", "source"]
    _header(ws, cols)
    for r in conn.execute(f"SELECT {','.join(cols)} FROM prices ORDER BY capture_date, sector, horizon, dep_time"):
        ws.append([r[c] for c in cols])
    _autosize(ws)

    # Sheet 2: curated view (exactly what the dashboard tables show)
    from render_dashboard import load
    entries_by_sector, _dates, _tb, _meta = load()
    ws2 = wb.create_sheet("Tracked flights")
    scols = ["capture_date", "sector", "airline", "row", "horizon",
             "target_date", "dep_time", "flight_number", "price"]
    _header(ws2, scols)
    name_by = {s["sector"]: s["name"] for s in config.SECTORS}
    flat = []
    for sector, groups in entries_by_sector.items():
        for airline, entries in groups:
            for e in entries:
                for horizon, pts in e["byh"].items():
                    for p in pts:
                        flat.append([p["date"], name_by.get(sector, sector), airline,
                                     e["label"], horizon, p["target_date"],
                                     p["dep_time"], p["flight_number"], p["price"]])
    flat.sort(key=lambda r: (r[0], r[1], r[2], str(r[6]), r[4]))
    for row in flat:
        ws2.append(row)
    _autosize(ws2)

    conn.close()
    wb.save(path)
    return path


def build_aviation_excel(path=AVIATION_XLSX):
    """Export civil-aviation metrics from aviation.db (long table + pivots)."""
    import os
    wb = Workbook()

    # Sheet 1: full long/tidy history
    ws = wb.active
    ws.title = "Daily long"
    cols = ["capture_date", "report_date", "section", "item",
            "value_num", "value_raw", "unit", "source"]
    _header(ws, cols)

    rows = []
    if os.path.exists(AVIATION_DB):
        conn = sqlite3.connect(AVIATION_DB)
        conn.row_factory = sqlite3.Row
        try:
            rows = conn.execute(
                f"SELECT {','.join(cols)} FROM aviation ORDER BY report_date, section, item"
            ).fetchall()
        except sqlite3.OperationalError:
            rows = []
        conn.close()
    for r in rows:
        ws.append([r[c] for c in cols])
    _autosize(ws)

    # Helper to build a pivot: report_date (rows) x column_key
    def pivot(sheet_name, predicate, col_key, value_key="value_num"):
        ws2 = wb.create_sheet(sheet_name)
        dates, colnames, cell = [], [], {}
        for r in rows:
            if not predicate(r):
                continue
            d, c = r["report_date"], col_key(r)
            if d not in dates:
                dates.append(d)
            if c not in colnames:
                colnames.append(c)
            cell[(d, c)] = r[value_key]
        _header(ws2, ["report_date"] + colnames)
        for d in dates:
            ws2.append([d] + [cell.get((d, c)) for c in colnames])
        _autosize(ws2)

    # Sheet 2: per-airline Passenger Load Factor over time
    pivot("Load factor",
          lambda r: r["section"] == "Passenger Load Factor",
          lambda r: r["item"])

    # Sheet 3: key traffic counts over time (domestic + international departures/pax)
    keep = {("Domestic traffic", "Departing flights"), ("Domestic traffic", "Departing Pax"),
            ("International traffic", "Departing flights"), ("International traffic", "Departing Pax")}
    scope = {"Domestic traffic": "Dom", "International traffic": "Int"}
    pivot("Traffic",
          lambda r: (r["section"], r["item"]) in keep,
          lambda r: f"{scope[r['section']]} {r['item']}")

    wb.save(path)
    return path


if __name__ == "__main__":
    print("Wrote", build_fares_excel())
    print("Wrote", build_aviation_excel())
